#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
vivo 工作流模块
功能：
  1. 扫描软著文件夹 → OCR 提取信息 → 生成 game.xlsx
  2. 读取 game.xlsx + user.xlsx → 调用 vivo 广告联盟 API → 生成渠广 txt 文件

vivo API 封装来自 vivomedia.py，已去除 pyppeteer 浏览器依赖，
仅保留纯 HTTP API 调用部分（不含创建新游戏的浏览器操作）。
"""


# core/vivo_workflow.py 中添加浏览器创建游戏的类

import os
import re
import hashlib
import base64
import time
import logging
import threading
from pathlib import Path
from typing import Dict, List, Optional, Callable

import requests

logger = logging.getLogger("softdoc_generator")


import asyncio
from pyppeteer import launch
from pypinyin import lazy_pinyin, Style

class VivoBrowserCreator:
    """使用浏览器创建 vivo 游戏并获取参数"""
    
    def __init__(self, account_name: str, password: str, secret_key: str, 
                 chrome_path: str = None, log_cb: Callable = None):
        self.account_name = account_name
        self.password = password
        self.secret_key = secret_key
        self.chrome_path = chrome_path or self._get_chrome_path()
        self.log_cb = log_cb
        self.browser = None
        self.page = None
    
    def _get_chrome_path(self) -> str:
        """获取 Chrome 可执行文件路径"""
        # 当前目录下的 chrome-win
        local_chrome = os.path.join(os.path.dirname(os.path.abspath(__file__)), "chrome-win", "chrome.exe")
        if os.path.exists(local_chrome):
            return local_chrome
        # 常见安装路径
        common_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe"),
        ]
        for path in common_paths:
            if os.path.exists(path):
                return path
        return None
    
    def _log(self, msg: str):
        logger.info(msg)
        if self.log_cb:
            self.log_cb(msg)
        
    async def _init_browser(self):
        """初始化浏览器"""
        self._log("正在启动浏览器...")
        
        # 设置环境变量，禁用信号处理（解决子线程问题）
        os.environ['PYPPETEER_DISABLE_SIGNAL_HANDLERS'] = '1'
        
        try:
            self.browser = await launch(
                headless=False,
                autoClose=False,
                executablePath=self.chrome_path,
                userDataDir=f"vivo_browser_{self.account_name}",
                args=[
                    "--disable-infobars", 
                    "--no-sandbox", 
                    "--disable-setuid-sandbox",
                    '--window-size=1200,1000', 
                    '--disable-dev-shm-usage',
                    '--disable-gpu',  # 添加
                    '--disable-software-rasterizer',  # 添加
                ],
                handleSIGINT=False,  # 禁用信号处理
                handleSIGTERM=False,  # 禁用信号处理
                handleSIGHUP=False,  # 禁用信号处理
            )
            self.page = await self.browser.newPage()
            await self.page.setViewport({"width": 1200, "height": 900})
            self._log("浏览器启动成功")
        except Exception as e:
            self._log(f"浏览器启动失败: {e}")
            raise
    
    async def _login(self):
        """登录 vivo 开发者后台"""
        url = "https://dev.vivo.com.cn/quickGame/create"
        await self.page.goto(url)
        await asyncio.sleep(3)
        
        current_url = self.page.url
        if current_url != url:
            self._log("需要登录，正在填写账号密码...")
            # 等待输入框加载
            await self.page.waitForSelector('input[type="text"]')
            await self.page.waitForSelector('input[type="password"]')
            
            # 填写用户名
            await self.page.type('input[type="text"]', self.account_name)
            await asyncio.sleep(1)
            
            # 填写密码
            await self.page.type('input[type="password"]', self.password)
            await asyncio.sleep(1)
            
            # 点击登录按钮
            submit_button = await self.page.querySelector('button[type="button"]')
            if submit_button:
                await submit_button.click()
                await asyncio.sleep(5)
            self._log("登录成功")
        else:
            self._log("已登录状态")
    
    def _get_pinyin_initials(self, text: str) -> str:
        """获取拼音首字母"""
        if not text:
            return "unknown"
        initials = ''.join([lazy_pinyin(char, style=Style.FIRST_LETTER)[0] for char in text if char.strip()])
        return initials.lower() if initials else "unknown"
    
    def _generate_package_name(self, company_name: str, game_name: str) -> str:
        """自动生成包名"""
        company_initials = self._get_pinyin_initials(company_name)
        game_initials = self._get_pinyin_initials(game_name)
        return f"com.{company_initials}.{game_initials}.vivominigame"
                
    async def create_or_get_game(self, game_name: str, company_name: str, 
                                existing_package: str = None) -> dict:
        """
        创建或获取游戏信息（完全参照 vivomedia.py 的 trans_vivo_rpk 函数）
        返回: {
            'game_name': 游戏名,
            'package_name': 包名,
            'app_key': App-key,
            'app_id': App-ID,
            'cp_id': Cp-ID,
            'media_id': 媒体ID (从API获取)
        }
        """
        try:
            # 1. 初始化浏览器并登录
            await self._init_browser()
            await self._login()
            
            # 2. 确定包名
            package_name = existing_package
            if not package_name or not package_name.strip():
                package_name = self._generate_package_name(company_name, game_name)
                self._log(f"自动生成包名: {package_name}")
            
            # 3. 进入创建页面
            url = "https://dev.vivo.com.cn/quickGame/create"
            await self.page.goto(url)
            await asyncio.sleep(5)
            
            # 4. 填写游戏信息并创建（完全参照 vivomedia.py 第 284-380 行）
            nn = 0
            is_stop_pkg = False
            is_stop_name = False
            
            while True:
                try:
                    if nn > 0:
                        url_text = 'https://dev.vivo.com.cn/quickGame/create'
                        await self.page.goto(url_text)
                        await asyncio.sleep(5)
                    
                    input_text = ".createquickgame input"
                    await self.page.waitForSelector(input_text)
                    input_els = await self.page.querySelectorAll(input_text)
                    
                    if nn == 0:
                        current_package = package_name
                        current_game_name = game_name
                    
                    await self.page.type("input[type='text'][placeholder='由英文字母或数字组成']", current_package)
                    await self.page.type("input[type='text'][placeholder='与RPK内的游戏名称保持一致']", current_game_name)
                    await asyncio.sleep(1)
                    
                    span_text = ".createquickgame span"
                    span_els = await self.page.querySelectorAll(span_text)
                    
                    response_promise = asyncio.ensure_future(
                        self.page.waitForResponse(lambda response: 'quickGame/verify-rpk-name' in response.url
                                                and response.request.method == 'POST')
                    )
                    response_promise_pkg = asyncio.ensure_future(
                        self.page.waitForResponse(lambda response: 'quickGame/verify-rpk-pkg' in response.url
                                                and response.request.method == 'POST')
                    )
                    
                    await self._click_element(span_els, '立即创建')
                    
                    response_text = {}
                    try:
                        self._log("正在等待名称验证响应")
                        specific_response = await asyncio.wait_for(response_promise, timeout=5.0)
                        response_text = await specific_response.json()
                        self._log(f"名称响应: {response_text}")
                    except asyncio.TimeoutError:
                        self._log("等待名称响应超时")
                        response_promise.cancel()
                    except Exception as e:
                        self._log(f"获取名称响应时出错: {e}")
                        response_promise.cancel()
                    
                    if 'code' in response_text:
                        if response_text['code'] == 0:
                            is_stop_name = True
                            self._log("名称正常")
                        else:
                            self._log(f"名称错误: {response_text.get('msg', '')}")
                            current_game_name = current_game_name + "手机游戏软件"
                    
                    response_text_pkg = {}
                    try:
                        self._log("正在等待包名验证响应")
                        specific_response_pkg = await asyncio.wait_for(response_promise_pkg, timeout=5.0)
                        response_text_pkg = await specific_response_pkg.json()
                        self._log(f"包名响应: {response_text_pkg}")
                    except asyncio.TimeoutError:
                        self._log("等待包名响应超时")
                        response_promise_pkg.cancel()
                    except Exception as e:
                        self._log(f"获取包名响应时出错: {e}")
                        response_promise_pkg.cancel()
                    
                    if 'code' in response_text_pkg:
                        if response_text_pkg['code'] == 0:
                            is_stop_pkg = True
                            self._log("包名正常")
                        else:
                            self._log(f"包名错误: {response_text_pkg.get('msg', '')}")
                            package_name_arr = current_package.split(".")
                            package_name_arr[2] = package_name_arr[2] + "r"
                            current_package = ".".join(package_name_arr)
                    
                    nn += 1
                    if (is_stop_name and is_stop_pkg) or nn > 1:
                        break
                    await asyncio.sleep(3)
                except Exception as err:
                    self._log(f"创建循环异常: {err}")
                    nn += 1
            
            # 更新最终的游戏名和包名
            game_name = current_game_name
            package_name = current_package
            
            # 5. 点击同意协议并进入详情页（完全参照 vivomedia.py 第 385-395 行）
            try:
                agree_text = ".agreeBt"
                await self.page.waitForSelector(agree_text)
                detail_text = ".breadcrumb-container .el-breadcrumb__inner"
                await self.page.waitForSelector(detail_text)
                detail_els = await self.page.querySelectorAll(detail_text)
                self._log("点击详情页")
                await self._click_element(detail_els, '小游戏详情页')
                await asyncio.sleep(1)
            except Exception as err:
                self._log(f"进入详情页失败: {err}")
            
            # 6. 等待详情页加载并获取 app 信息（完全参照 vivomedia.py 第 397-412 行）
            await self.page.waitForSelector(".title-warp")
            app_text = ".app-detail-wrap .info-wrap .app-back-warp div"
            await self.page.waitForSelector(app_text)
            detail_div_els = await self.page.querySelectorAll(app_text)
            
            app_key = ""
            app_id = ""
            cp_id = ""
            
            for detail_div_el in detail_div_els:
                text = await self.page.evaluate("(el) => el.textContent", detail_div_el)
                self._log(f"详情页内容: {text}")
                if "App-key" in text:
                    app_key = text.replace("App-key", "").strip()
                if "App-ID" in text:
                    app_id = text.replace("App-ID", "").strip()
                if "Cp-ID" in text:
                    cp_id = text.replace("Cp-ID", "").strip()
            
            self._log(f"app_key: {app_key}, app_id: {app_id}, cp_id: {cp_id}")
            
            # 7. 通过 API 获取 media_id
            media_id = await self._get_media_id_via_api(package_name, game_name)
            
            result = {
                'game_name': game_name,
                'package_name': package_name,
                'app_key': app_key,
                'app_id': app_id,
                'cp_id': cp_id,
                'media_id': media_id,
            }
            
            self._log(f"游戏创建/获取成功: {result}")
            return result
            
        except Exception as e:
            self._log(f"浏览器创建游戏失败: {e}")
            import traceback
            traceback.print_exc()
            return {
                'game_name': game_name,
                'package_name': existing_package or self._generate_package_name(company_name, game_name),
                'app_key': '',
                'app_id': '',
                'cp_id': '',
                'media_id': '',
            }
        finally:
            await self._close()


    async def _click_element(self, elements, click_text):
        """点击包含指定文本的元素（完全参照 vivomedia.py）"""
        for element in elements:
            text = await self.page.evaluate('(el) => el.textContent', element)
            if text.strip() == click_text:
                await element.click()
                break


    async def _get_app_info(self) -> dict:
        """从详情页获取 app_key, app_id, cp_id"""
        result = {'app_key': '', 'app_id': '', 'cp_id': ''}
        
        try:
            # 等待详情页信息区域加载
            await self.page.waitForSelector(".app-detail-wrap .info-wrap .app-back-warp div", timeout=15000)
            
            # 获取所有信息 div
            detail_div_els = await self.page.querySelectorAll(".app-detail-wrap .info-wrap .app-back-warp div")
            
            self._log(f"找到 {len(detail_div_els)} 个信息元素")
            
            for div_el in detail_div_els:
                text = await self.page.evaluate("(el) => el.textContent", div_el)
                
                if "App-key" in text:
                    result['app_key'] = text.replace("App-key", "").strip()
                    self._log(f"✓ 获取到 App-key: {result['app_key'][:20]}...")
                elif "App-ID" in text:
                    result['app_id'] = text.replace("App-ID", "").strip()
                    self._log(f"✓ 获取到 App-ID: {result['app_id']}")
                elif "Cp-ID" in text:
                    result['cp_id'] = text.replace("Cp-ID", "").strip()
                    self._log(f"✓ 获取到 Cp-ID: {result['cp_id']}")
            
            # 备用方案：如果选择器没找到，尝试从整个页面提取
            if not any(result.values()):
                self._log("尝试备用方式获取 app 信息...")
                html = await self.page.content()
                import re
                
                # 提取 App-key
                match = re.search(r'App-key[：:\s]*([A-Za-z0-9]+)', html)
                if match:
                    result['app_key'] = match.group(1)
                    self._log(f"备用获取 App-key: {result['app_key'][:20]}...")
                
                # 提取 App-ID
                match = re.search(r'App-ID[：:\s]*([A-Za-z0-9]+)', html)
                if match:
                    result['app_id'] = match.group(1)
                    self._log(f"备用获取 App-ID: {result['app_id']}")
                
                # 提取 Cp-ID
                match = re.search(r'Cp-ID[：:\s]*([A-Za-z0-9]+)', html)
                if match:
                    result['cp_id'] = match.group(1)
                    self._log(f"备用获取 Cp-ID: {result['cp_id']}")
                        
        except Exception as e:
            self._log(f"获取 App 信息失败: {e}")
        
        return result


    async def _get_media_id_via_api(self, package_name: str, game_name: str) -> str:
        """通过 API 获取 media_id"""
        from .vivo_workflow import VivoAdAPI
        
        api = VivoAdAPI(self.account_name, self.secret_key)
        media_id = api.get_or_create_media_id(package_name, game_name, log_cb=self._log)
        return media_id or ""
    
    async def _wait_for_detail_page(self, timeout: int = 30):
        """等待进入详情页"""
        start = asyncio.get_event_loop().time()
        while (asyncio.get_event_loop().time() - start) < timeout:
            current_url = self.page.url
            if 'quickGame/detail' in current_url or 'app-detail-wrap' in current_url:
                self._log("已进入详情页")
                await asyncio.sleep(2)
                return
            await asyncio.sleep(1)
        self._log("未检测到详情页，继续尝试获取信息")
        
  
    
 
    async def _close(self):
        """关闭浏览器"""
        try:
            if self.page and not self.page.isClosed():
                await self.page.close()
            if self.browser:
                await self.browser.close()
        except Exception as e:
            self._log(f"关闭浏览器时出错: {e}")


# core/vivo_workflow.py 中修改 create_game_with_browser 函数

def create_game_with_browser(account_name: str, password: str, secret_key: str,
                              game_name: str, company_name: str, package_name: str = None,
                              log_cb: Callable = None) -> dict:
    """
    同步包装器，调用浏览器创建游戏
    使用新的事件循环，避免信号问题
    """
    import asyncio
    import sys
    
    async def _async_create():
        creator = VivoBrowserCreator(account_name, password, secret_key, log_cb=log_cb)
        return await creator.create_or_get_game(game_name, company_name, package_name)
    
    # 创建新的事件循环
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        result = loop.run_until_complete(_async_create())
        return result
    except Exception as e:
        logger.error(f"create_game_with_browser 失败: {e}")
        import traceback
        traceback.print_exc()
        # 返回基本信息
        return {
            'game_name': game_name,
            'package_name': package_name or '',
            'app_key': '',
            'app_id': '',
            'cp_id': '',
            'media_id': '',
        }
    finally:
        try:
            loop.close()
        except:
            pass



# ──────────────────────────────────────────────────────────────────── #
#  工具函数
# ──────────────────────────────────────────────────────────────────── #

def _send_request(url: str, data=None, method: str = 'GET',
                  headers=None, max_retries: int = 3) -> Optional[dict]:
    """通用 HTTP 请求（带重试）"""
    retries = 0
    delay = 1.0
    while retries <= max_retries:
        try:
            kwargs = {
                'url': url,
                'timeout': 15,
                'headers': headers or {}
            }
            if method.upper() == 'POST':
                kwargs['json'] = data
                response = requests.post(**kwargs)
            else:
                kwargs['params'] = data
                response = requests.get(**kwargs)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.HTTPError as e:
            if e.response and 400 <= e.response.status_code < 500:
                logger.warning(f"客户端错误 {e.response.status_code}，不重试")
                return None
        except requests.exceptions.RequestException as e:
            logger.warning(f"请求异常 (尝试 {retries+1}): {e}")
        if retries < max_retries:
            time.sleep(delay)
            delay *= 2
        retries += 1
    logger.error(f"请求失败，已重试 {max_retries} 次: {url}")
    return None


def _create_token(account_name: str, secret_key: str) -> str:
    """生成 vivo 广告联盟 API Token"""
    timestamp = int(round(time.time() * 1000))
    sign = hashlib.sha256(
        (account_name + secret_key + str(timestamp)).encode('utf-8')
    ).hexdigest()
    plain = f"{timestamp},{sign}"
    return base64.b64encode(plain.encode('utf-8')).decode('utf-8')


# ──────────────────────────────────────────────────────────────────── #
#  vivo 广告联盟 API
# ──────────────────────────────────────────────────────────────────── #

class VivoAdAPI:
    """vivo 广告联盟 API 封装（纯 HTTP，无浏览器）"""

    BASE_URL = "https://adnet.vivo.com.cn"

    def __init__(self, account_name: str, secret_key: str):
        self.account_name = account_name
        self.secret_key = secret_key

    def _headers(self) -> dict:
        token = _create_token(self.account_name, self.secret_key)
        return {'token': token}

    def _params_base(self) -> dict:
        return {'accountName': self.account_name}

    # ---- 媒体查询 ----

    def get_media_by_package(self, package_name: str) -> Optional[dict]:
        """按包名查询媒体"""
        url = self.BASE_URL + "/api/open/media/list"
        params = {**self._params_base(), 'packageName': package_name}
        return _send_request(url, data=params, method='GET', headers=self._headers())

    def get_media_by_name(self, media_name: str) -> Optional[dict]:
        """按媒体名称查询媒体"""
        url = self.BASE_URL + "/api/open/media/list"
        params = {**self._params_base(), 'mediaName': media_name}
        return _send_request(url, data=params, method='GET', headers=self._headers())

    def create_media(self, package_name: str, media_name: str,
                     keyword: str = '游戏') -> Optional[dict]:
        """创建媒体"""
        url = self.BASE_URL + "/api/open/media/add"
        data = {
            **self._params_base(),
            'packageName': package_name,
            'mediaName': media_name,
            'keywords': [keyword],
        }
        return _send_request(url, data=data, method='POST', headers=self._headers())

    # ---- 广告位查询 / 创建 ----

    def get_positions_by_package(self, package_name: str) -> Optional[dict]:
        """按包名查询广告位列表"""
        url = self.BASE_URL + "/api/open/position/list"
        params = {**self._params_base(), 'packageName': package_name}
        return _send_request(url, data=params, method='GET', headers=self._headers())

    def create_position(self, media_id: str, pos_name: str,
                        pos_type: int, extra: dict = None) -> Optional[dict]:
        """
        创建广告位
        pos_type: 2=开屏, 3=banner, 4=插屏, 5=原生, 9=激励视频
        """
        url = self.BASE_URL + "/api/open/position/add"
        data = {
            **self._params_base(),
            'mediaId': media_id,
            'positionName': pos_name,
            'positionType': pos_type,
            'accessType': 1,   # API 接入
            'secondScene': 1,  # 通用
        }
        if extra:
            data.update(extra)
        return _send_request(url, data=data, method='POST', headers=self._headers())

    # ---- 高层封装：获取或创建媒体 ID ----

    def get_or_create_media_id(self, package_name: str, media_name: str,
                               log_cb: Callable = None) -> Optional[str]:
        """
        查询包名是否已有媒体。
        - 找到 → 返回 media_id
        - 没有 → 创建媒体 → 返回 media_id
        - API 未授权 / 失败 → 返回 None
        """
        def _log(msg):
            logger.info(msg)
            if log_cb:
                log_cb(msg)

        # 1. 按包名查询
        ret = self.get_media_by_package(package_name)
        if ret is None:
            _log(f"查询媒体失败（网络错误）: {package_name}")
            return None
        if ret.get('code') == 40001:
            _log("查询媒体接口未授权，请联系 vivo 开通 API 权限")
            return None
        if ret.get('code') == 0 and ret.get('data', {}).get('list'):
            media_id = ret['data']['list'][0]['mediaId']
            _log(f"查询到媒体: {media_id} ({package_name})")
            return media_id

        # 2. 按游戏名查询（防止包名不一致但游戏名一样）
        ret2 = self.get_media_by_name(media_name)
        if ret2 and ret2.get('code') == 0 and ret2.get('data', {}).get('list'):
            for item in ret2['data']['list']:
                if item.get('mediaType') == '快游戏':
                    media_id = item['mediaId']
                    _log(f"按游戏名查询到媒体: {media_id} ({media_name})")
                    return media_id

        # 3. 创建媒体
        _log(f"媒体不存在，正在创建: {media_name} ({package_name})")
        ret3 = self.create_media(package_name, media_name)
        if ret3 is None:
            _log("创建媒体失败（网络错误）")
            return None
        if ret3.get('code') == 40001:
            _log("创建媒体接口未授权，请联系 vivo 开通 API 权限")
            return None
        if ret3.get('code') == 2001:
            # 媒体名称重复，重新按包名查
            _log("媒体名称重复，重新查询...")
            ret4 = self.get_media_by_package(package_name)
            if ret4 and ret4.get('code') == 0 and ret4.get('data', {}).get('list'):
                media_id = ret4['data']['list'][0]['mediaId']
                _log(f"重新查询到媒体: {media_id}")
                return media_id
            return None
        if ret3.get('code') == 0:
            media_id = ret3['data']['mediaId']
            _log(f"创建媒体成功: {media_id}")
            return media_id

        _log(f"创建媒体失败: {ret3}")
        return None

    # ---- 高层封装：确保广告位存在 ----

    def ensure_positions(self, media_id: str, package_name: str,
                         ads_config: dict, game_name: str,
                         log_cb: Callable = None) -> List[dict]:
        """
        根据 ads_config 确保广告位都存在（不存在则创建）。
        返回：[{'positionName': ..., 'positionId': ...}, ...]

        ads_config 格式（与 vivomedia.py 一致）：
        {
          'open':      {'number': 1, 'text': 'H5-开屏'},
          'banner':    {'number': 1, 'text': 'H5-banner'},
          'reward':    {'number': 1, 'text': 'H5-激励视频'},
          'ori':       {'number': 1, 'text': 'H5-结算模版'},
          'ori_other': {'number': 1, 'text': 'H5-其它模版'},
        }
        """
        def _log(msg):
            logger.info(msg)
            if log_cb:
                log_cb(msg)

        # 查询已有广告位
        existing_names = set()
        ret = self.get_positions_by_package(package_name)
        if ret and ret.get('code') == 0 and ret.get('data', {}).get('list'):
            for item in ret['data']['list']:
                if '原生小卡' not in item.get('positionName', ''):
                    existing_names.add(item['positionName'])

        POS_TYPE_MAP = {
            'open': (2, {'orientation': 1, 'secondScene': 1}),
            'banner': (3, {'orientation': 1, 'secondScene': 1, 'renderStyle': [31, 33, 32]}),
            'reward': (9, {'secondScene': 1}),
            'ori': (5, {'secondScene': 1, 'renderType': 1, 'renderStyles': [5]}),
            'ori_other': (5, {'secondScene': 1, 'renderType': 1, 'renderStyles': [5]}),
        }

        ad_list = []
        for ad_key, pos_info in POS_TYPE_MAP.items():
            if ad_key not in ads_config:
                continue
            cfg = ads_config[ad_key]
            num = int(cfg.get('number', 1))
            text = cfg.get('text', ad_key)
            pos_type, extra = pos_info

            for i in range(num):
                suffix = str(i + 1) if num > 1 else ''
                name = game_name + text + suffix
                if name in existing_names:
                    _log(f"广告位已存在，跳过: {name}")
                    # 从已有列表找到 positionId
                    if ret and ret.get('data', {}).get('list'):
                        for item in ret['data']['list']:
                            if item.get('positionName') == name:
                                ad_list.append({
                                    'positionName': name,
                                    'positionId': item['positionId']
                                })
                    continue

                result = self.create_position(media_id, name, pos_type, extra)
                if result and result.get('code') == 0:
                    pos_id = result['data']['positionId']
                    ad_list.append({'positionName': name, 'positionId': pos_id})
                    _log(f"已创建广告位: {name} → {pos_id}")
                elif result and result.get('code') == 40001:
                    _log("创建广告位接口未授权")
                else:
                    _log(f"创建广告位失败: {name} → {result}")

        return ad_list


# ──────────────────────────────────────────────────────────────────── #
#  game.xlsx 生成
# ──────────────────────────────────────────────────────────────────── #

# core/vivo_workflow.py 中修改 generate_game_xlsx 函数

# core/vivo_workflow.py 中修改 generate_game_xlsx 函数

# core/vivo_workflow.py 中修改 generate_game_xlsx 函数

def generate_game_xlsx(softdoc_infos: List[dict], output_path: str) -> bool:
    """
    根据软著 OCR 信息列表生成 game.xlsx。
    包名会自动生成（不需要用户补充）。
    """
    try:
        import openpyxl
        from pypinyin import lazy_pinyin, Style
        
        print(f"\n【DEBUG generate_game_xlsx】")
        print(f"  原始 output_path: {output_path}")
        print(f"  类型: {type(output_path)}")
        
        # ---- 关键修复：处理路径 ----
        output_path = os.path.normpath(output_path)  # 标准化路径
        print(f"  标准化后: {output_path}")
        
        # 检查路径是否包含非法字符
        illegal_chars = r'<>:"|?*'
        for char in illegal_chars:
            if char in output_path:
                print(f"  警告: 路径包含非法字符 '{char}'")
        
        # 情况1：如果路径是一个已存在的目录，则在该目录下创建 game.xlsx
        if os.path.isdir(output_path):
            output_path = os.path.join(output_path, "game.xlsx")
            print(f"  是目录，拼接后: {output_path}")
        # 情况2：如果路径以 .xlsx 结尾，直接使用
        elif output_path.lower().endswith('.xlsx'):
            print(f"  已是 .xlsx 文件")
        # 情况3：其他情况，添加 .xlsx
        else:
            output_path += ".xlsx"
            print(f"  添加扩展名: {output_path}")
        
        # 检查路径是否包含双反斜杠或奇怪字符
        print(f"  最终路径: {output_path}")
        
        # 确保目标目录存在
        target_dir = os.path.dirname(output_path)
        print(f"  目标目录: {target_dir}")
        
        if target_dir:
            # 检查目录是否可写
            if os.path.exists(target_dir):
                is_writable = os.access(target_dir, os.W_OK)
                print(f"  目录存在，是否可写: {is_writable}")
                if not is_writable:
                    logger.error(f"目录不可写: {target_dir}")
                    return False
            else:
                print(f"  目录不存在，尝试创建: {target_dir}")
                os.makedirs(target_dir, exist_ok=True)
                print(f"  目录创建成功")
        
        # 检查文件是否已存在且被占用
        if os.path.exists(output_path):
            print(f"  文件已存在: {output_path}")
            try:
                # 尝试以只读方式打开，检查是否被占用
                with open(output_path, 'rb') as f:
                    pass
                print(f"  文件未被占用，将覆盖")
            except PermissionError as e:
                print(f"  文件被占用，无法写入: {e}")
                logger.error(f"文件被占用: {output_path}, 请关闭该文件后重试")
                return False
        
        def get_pinyin_initials(text: str) -> str:
            """获取拼音首字母（小写）"""
            if not text:
                return "unknown"
            initials = ''.join([lazy_pinyin(char, style=Style.FIRST_LETTER)[0] for char in text if char.strip()])
            return initials.lower() if initials else "unknown"
        
        print(f"  开始生成 Excel...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "游戏信息"
        headers = ['主体', '游戏名称', '包名', '登记号']
        ws.append(headers)
        
        for info in softdoc_infos:
            main_name = info.get('main_name', info.get('copyright_holder', ''))
            game_name = info.get('game_name', '')
            
            # 自动生成包名
            company_initials = get_pinyin_initials(main_name)
            game_initials = get_pinyin_initials(game_name)
            auto_package = f"com.{company_initials}.{game_initials}.vivominigame"
            
            original_package = info.get('package_name', '')
            package_name = original_package if original_package and original_package.strip() else auto_package
            
            ws.append([
                main_name,
                game_name,
                package_name,
                info.get('registration_number', ''),
            ])
        
        print(f"  保存文件: {output_path}")
        wb.save(output_path)
        print(f"  保存成功!")
        
        logger.info(f"game.xlsx 已生成: {output_path}，共 {len(softdoc_infos)} 行")
        return True
    except PermissionError as e:
        logger.error(f"权限错误: {e}")
        print(f"【DEBUG】PermissionError: {e}")
        print(f"【DEBUG】问题路径: {output_path}")
        # 尝试用临时文件测试写入权限
        test_path = os.path.join(target_dir, "_test_write.tmp")
        try:
            with open(test_path, 'w') as f:
                f.write("test")
            os.remove(test_path)
            print(f"【DEBUG】目录写入测试成功，问题可能出在文件被占用")
        except Exception as test_e:
            print(f"【DEBUG】目录写入测试失败: {test_e}")
        return False
    except Exception as e:
        logger.error(f"生成 game.xlsx 失败: {e}")
        print(f"【DEBUG】Exception: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return False


# ──────────────────────────────────────────────────────────────────── #
#  渠广 txt 生成
# ──────────────────────────────────────────────────────────────────── #

def generate_qg_txt(output_dir: str,
                    game_name: str,
                    package_name: str,
                    app_key: str,
                    app_id: str,
                    cp_id: str,
                    full_company_name: str,
                    media_id: str,
                    ad_list: List[dict]) -> str:
    """
    生成渠广 txt 文件，格式与 vivomedia.py 一致。
    返回生成的文件路径。
    """
    os.makedirs(output_dir, exist_ok=True)
    filename = os.path.join(output_dir, f"{game_name}_vivo小游戏渠广.txt")

    lines = [
        game_name,
        f"包名：{package_name}",
        f"App-key：{app_key}",
        f"App-ID：{app_id}",
        f"Cp-ID：{cp_id}",
        "",
        full_company_name,
        "",
        game_name,
        f"MediaID：{media_id}",
        "",
    ]
    for ad in ad_list:
        lines.append(ad['positionName'])
        lines.append(f"posID：{ad['positionId']}")
    lines += [
        "",
        package_name.replace(".vivominigame", ""),
        "",
        "VIVOCP",
        "VIVOH5",
    ]

    with open(filename, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines) + '\n')

    logger.info(f"渠广 txt 已生成: {filename}")
    return filename


# ──────────────────────────────────────────────────────────────────── #
#  完整工作流
# ──────────────────────────────────────────────────────────────────── #

class VivoWorkflow:
    """
    整合「软著→game.xlsx→渠广 txt」的完整工作流。
    设计为在后台线程中运行，通过 log_callback 实时回调日志到 GUI。
    """

    DEFAULT_ADS_CONFIG = {
        'open':      {'number': 1, 'text': 'H5-开屏'},
        'banner':    {'number': 1, 'text': 'H5-banner'},
        'reward':    {'number': 1, 'text': 'H5-激励视频'},
        'ori':       {'number': 1, 'text': 'H5-结算模版'},
        'ori_other': {'number': 1, 'text': 'H5-其它模版'},
    }

    def __init__(self, config=None):
        self.config = config
        self._stop_event = threading.Event()

    def stop(self):
        """请求中止工作流"""
        self._stop_event.set()

    def reset(self):
        self._stop_event.clear()

    # ---- Step 1：软著 → 提取信息列表 ----

    # core/vivo_workflow.py 中修改 scan_softdoc_folder 方法

    # core/vivo_workflow.py 中修改 scan_softdoc_folder 方法

    # core/vivo_workflow.py 中修改 scan_softdoc_folder 方法

    # core/vivo_workflow.py 中修改 scan_softdoc_folder 方法

    def scan_softdoc_folder(self, softdoc_dir: str,
                            api_ocr=None,
                            log_cb: Callable = None) -> List[dict]:
        """
        扫描软著文件夹，OCR 提取所有软著信息。
        对文件夹内的每个 PDF/图片文件单独解析。
        返回 softdoc_infos 列表。
        """
        def _log(msg):
            logger.info(msg)
            if log_cb:
                log_cb(msg)

        from .softdoc_parser import SoftDocParser

        if not os.path.isdir(softdoc_dir):
            _log(f"软著文件夹不存在: {softdoc_dir}")
            return []

        # 收集所有软著文件
        supported_exts = ('.pdf', '.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp')
        all_files = []
        for f in os.listdir(softdoc_dir):
            if os.path.splitext(f.lower())[1] in supported_exts:
                all_files.append(os.path.join(softdoc_dir, f))
        all_files.sort()
        _log(f"找到 {len(all_files)} 个软著文件")

        if not all_files:
            _log("软著文件夹为空，无可处理文件")
            return []

        # 创建 parser
        parser = SoftDocParser(
            config=self.config,
            external_api_ocr=api_ocr
        )

        infos = []
        for i, fpath in enumerate(all_files, 1):
            if self._stop_event.is_set():
                _log("已中止")
                break
                
            fname = os.path.basename(fpath)
            _log(f"\n[{i}/{len(all_files)}] 解析: {fname}")
            
            try:
                # 对每个文件单独调用 parse_file
                result = parser.parse_file(fpath)
                
                software_name = result.get('software_name', '')
                copyright_holder = result.get('copyright_holder', '')
                registration_number = result.get('registration_number', '')
                
                if not software_name:
                    _log(f"  ⚠ 未提取到软件名称，跳过: {fname}")
                    continue

                # 从 software_name 中提取游戏名（使用 softdoc_parser 的 _extract_soft_info 已经处理过）
                import re
                game_name = software_name
                
                # 优先从《》中提取
                bracket_match = re.search(r'《([^》]+)》', software_name)
                if bracket_match:
                    game_name = bracket_match.group(1).strip()
                else:
                    # 去掉常见后缀
                    suffixes = ['手机游戏软件$', '游戏软件$', '软件$', 'V\d+\.\d+$']
                    for suffix in suffixes:
                        game_name = re.sub(suffix, '', game_name).strip()
                
                # 过滤无效的游戏名
                invalid_names = ['计算机软件保护条例', '计算机软件', '软件保护条例']
                if game_name in invalid_names or len(game_name) < 2:
                    _log(f"  ⚠ 游戏名无效: '{game_name}'，尝试从文件名提取...")
                    # 从文件名提取（去掉扩展名和常见后缀）
                    name_from_file = os.path.splitext(fname)[0]
                    for suffix in ['手机游戏软件', '游戏软件']:
                        name_from_file = name_from_file.replace(suffix, '')

                    # 从 copyright_holder 提取公司简称，用于清理文件名中拼进去的公司名
                    # 例如: copyright_holder="深圳快乐游科技有限公司" → 简称="快乐游"
                    company_abbrev = ''
                    if copyright_holder:
                        _comp = re.sub(r'(有限公司|科技有限公司|股份公司|网络科技)$', '', copyright_holder)
                        _comp = _comp.strip()
                        # 去掉省市区县前缀，取最后 2-6 个字符作为公司简称
                        # "深圳快乐游" → "快乐游"，"深圳王子星核" → "王子星核"
                        # 先去掉常见前缀（2-4字）
                        _prefixes = ['深圳', '广州', '北京', '上海', '杭州', '成都', '武汉', '南京', '西安']
                        for _p in _prefixes:
                            if _comp.startswith(_p):
                                _comp = _comp[len(_p):]
                                break
                        # 取最后 6 个字作为简称
                        company_abbrev = _comp[-6:].strip() if len(_comp) >= 2 else _comp

                    # 清理文件名中拼进去的公司名
                    if company_abbrev and company_abbrev in name_from_file:
                        name_from_file = name_from_file.replace(company_abbrev, '')
                        _log(f"  从文件名清理公司名: 简称='{company_abbrev}'")

                    # 去掉剩余的零散"软件"/"科技"等残留
                    for _rem in ['软件', '科技']:
                        name_from_file = name_from_file.replace(_rem, '')

                    name_from_file = re.sub(r'\s+', ' ', name_from_file).strip()
                    game_name = name_from_file
                    _log(f"  从文件名提取: {game_name}")

                info = {
                    'game_name': game_name,
                    'software_name': software_name,
                    'copyright_holder': copyright_holder,
                    'main_name': copyright_holder,
                    'registration_number': registration_number,
                    'package_name': '',
                    'source_file': fname,
                }
                infos.append(info)
                _log(f"  ✓ 游戏名: {game_name}")
                _log(f"    软件名: {software_name}")
                _log(f"    著作权人: {copyright_holder}")
                _log(f"    登记号: {registration_number}")
                
            except Exception as e:
                _log(f"  ✗ 解析失败: {fname} - {e}")
                import traceback
                traceback.print_exc()

        _log(f"\n扫描完成，共提取 {len(infos)} 条记录")
        return infos

    # ---- Step 2：game.xlsx → 渠广 txt ----

    # core/vivo_workflow.py 中完整的 generate_qg_files 方法

    def generate_qg_files(self,
                        game_xlsx_path: str,
                        user_xlsx_path: str,
                        output_dir: str,
                        ads_config: dict = None,
                        log_cb: Callable = None,
                        use_browser: bool = True) -> List[str]:
        """
        读取 game.xlsx 和 user.xlsx，调用 vivo API 生成渠广 txt 文件。
        返回生成的文件路径列表。
        
        Args:
            game_xlsx_path: game.xlsx 文件路径或目录
            user_xlsx_path: user.xlsx 文件路径或目录
            output_dir: 渠广文件输出目录
            ads_config: 广告位配置
            log_cb: 日志回调函数
            use_browser: 是否使用浏览器创建游戏（获取 app_key/app_id/cp_id）
        """
        def _log(msg):
            logger.info(msg)
            if log_cb:
                log_cb(msg)

        from openpyxl import load_workbook
        from .vivo_workflow import VivoAdAPI, generate_qg_txt

        if ads_config is None:
            ads_config = self.DEFAULT_ADS_CONFIG

        # ---- 处理 game.xlsx 路径 ----
        game_xlsx_path = os.path.normpath(game_xlsx_path)
        if os.path.isdir(game_xlsx_path):
            game_xlsx_path = os.path.join(game_xlsx_path, "game.xlsx")
            _log(f"game.xlsx 路径自动补全: {game_xlsx_path}")
        
        if not os.path.exists(game_xlsx_path):
            _log(f"❌ game.xlsx 文件不存在: {game_xlsx_path}")
            return []
        
        # ---- 处理 user.xlsx 路径 ----
        user_xlsx_path = os.path.normpath(user_xlsx_path)
        if os.path.isdir(user_xlsx_path):
            user_xlsx_path = os.path.join(user_xlsx_path, "user.xlsx")
            _log(f"user.xlsx 路径自动补全: {user_xlsx_path}")
        
        if not os.path.exists(user_xlsx_path):
            _log(f"❌ user.xlsx 文件不存在: {user_xlsx_path}")
            return []

        # ---- 读取 user.xlsx（使用 openpyxl）----
        _log(f"读取账号信息: {user_xlsx_path}")
        try:
            wb = load_workbook(user_xlsx_path)
            ws = wb.active
            
            user_map: Dict[str, dict] = {}
            # 假设表头: name, user, pwd, secret_key, full_name
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row or not row[0]:
                    continue
                    
                name = str(row[0]).strip() if row[0] else ''
                username = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                password = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                secret_key = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                full_name = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                
                if name:
                    user_map[name] = {
                        'user': username,
                        'pwd': password,
                        'secret_key': secret_key,
                        'full_name': full_name,
                    }
                    _log(f"  加载账号: {name} -> {username}")
            
            _log(f"账号信息加载完成，共 {len(user_map)} 个主体: {list(user_map.keys())}")
            
        except Exception as e:
            _log(f"读取 user.xlsx 失败: {e}")
            import traceback
            traceback.print_exc()
            return []

        # ---- 读取 game.xlsx（使用 openpyxl）----
        _log(f"读取游戏信息: {game_xlsx_path}")
        try:
            wb = load_workbook(game_xlsx_path)
            ws = wb.active
            
            games = []
            # 假设表头: 主体, 游戏名称, 包名, 登记号
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row or not row[1]:  # 游戏名称不能为空
                    continue
                    
                main_name = str(row[0]).strip() if row[0] else ''
                game_name = str(row[1]).strip() if row[1] else ''
                package_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                registration_number = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                
                games.append({
                    'main_name': main_name,
                    'game_name': game_name,
                    'package_name': package_name,
                    'registration_number': registration_number,
                })
                _log(f"  加载游戏: {game_name} (主体: {main_name})")
            
            _log(f"游戏信息加载完成，共 {len(games)} 个游戏")
            
        except Exception as e:
            _log(f"读取 game.xlsx 失败: {e}")
            import traceback
            traceback.print_exc()
            return []

        if not games:
            _log("❌ game.xlsx 中没有有效数据")
            return []

        # ---- 生成渠广文件 ----
        os.makedirs(output_dir, exist_ok=True)
        generated_files = []
        
        from pypinyin import lazy_pinyin, Style
        from .vivo_workflow import create_game_with_browser

        for idx, game in enumerate(games, 1):
            if self._stop_event.is_set():
                _log("已中止")
                break

            main_name = game['main_name']
            game_name = game['game_name']
            package_name = game['package_name']
            
            if not game_name:
                _log(f"跳过空行")
                continue

            _log(f"\n{'='*50}")
            _log(f"[{idx}/{len(games)}] 处理: {game_name}")
            _log(f"  主体: {main_name}")
            _log(f"  包名: {package_name if package_name else '(将自动生成)'}")

            # ---- 找账号 ----
            user_info = user_map.get(main_name)
            if not user_info:
                # 尝试模糊匹配
                for k, v in user_map.items():
                    if k in main_name or main_name in k:
                        user_info = v
                        _log(f"  模糊匹配账号: '{k}' 对应 '{main_name}'")
                        break
                        
            if not user_info:
                _log(f"  ⚠ 未找到主体 '{main_name}' 对应的账号，跳过")
                continue

            account_name = user_info['user']
            password = user_info.get('pwd', '')
            secret_key = user_info['secret_key']
            full_name = user_info['full_name']

            if not account_name or not secret_key:
                _log(f"  ⚠ 账号信息不完整（user/secret_key 为空），跳过")
                continue

            # ---- 初始化变量 ----
            app_key = ''
            app_id = ''
            cp_id = ''
            media_id = ''
            final_package_name = package_name

            # ---- 方案一：使用浏览器创建游戏（获取完整信息）----
            use_browser_fallback = True
            
            if use_browser:
                if not password:
                    _log(f"  ⚠ 浏览器模式需要密码，但 user.xlsx 中 password 为空，将使用 API 模式")
                else:
                    try:
                        _log(f"  🌐 使用浏览器创建/获取游戏...")
                        
                        game_result = create_game_with_browser(
                            account_name=account_name,
                            password=password,
                            secret_key=secret_key,
                            game_name=game_name,
                            company_name=main_name,
                            package_name=package_name if package_name else None,
                            log_cb=_log
                        )
                        
                        final_package_name = game_result.get('package_name', package_name)
                        app_key = game_result.get('app_key', '')
                        app_id = game_result.get('app_id', '')
                        cp_id = game_result.get('cp_id', '')
                        media_id = game_result.get('media_id', '')
                        
                        _log(f"  ✓ 浏览器创建完成:")
                        _log(f"      包名: {final_package_name}")
                        _log(f"      App-key: {app_key[:20] if app_key else '(空)'}...")
                        _log(f"      App-ID: {app_id}")
                        _log(f"      Cp-ID: {cp_id}")
                        _log(f"      MediaID: {media_id if media_id else '(待获取)'}")
                        use_browser_fallback = False
                        
                    except Exception as e:
                        _log(f"  ✗ 浏览器创建失败: {e}")
                        _log(f"  将回退到 API 模式")
                        use_browser_fallback = True

            # ---- 方案二：API 模式（或浏览器模式回退）----
            if use_browser_fallback:
                _log(f"  🔧 使用 API 模式获取/创建媒体...")
                
                api = VivoAdAPI(account_name, secret_key)
                
                # 获取或创建 media_id
                if not final_package_name:
                    # 自动生成包名
                    def get_initials(text: str) -> str:
                        if not text:
                            return "unknown"
                        initials = ''.join([lazy_pinyin(char, style=Style.FIRST_LETTER)[0] 
                                            for char in text if char.strip()])
                        return initials.lower() if initials else "unknown"
                    
                    company_initials = get_initials(main_name)
                    game_initials = get_initials(game_name)
                    final_package_name = f"com.{company_initials}.{game_initials}.vivominigame"
                    _log(f"  自动生成包名: {final_package_name}")
                
                # 获取 media_id
                media_id = api.get_or_create_media_id(final_package_name, game_name, log_cb=_log)
                if not media_id:
                    _log(f"  ✗ 无法获取 media_id，跳过: {game_name}")
                    continue
                
                _log(f"  ✓ MediaID: {media_id}")

            # ---- 创建/查询广告位 ----
            _log(f"  📢 处理广告位...")
            api = VivoAdAPI(account_name, secret_key)
            ad_list = api.ensure_positions(
                media_id, final_package_name, ads_config, game_name, log_cb=_log
            )
            
            _log(f"  广告位数量: {len(ad_list)}")
            for ad in ad_list:
                _log(f"      {ad['positionName']} -> {ad['positionId']}")

            # ---- 生成渠广 txt ----
            txt_path = generate_qg_txt(
                output_dir=output_dir,
                game_name=game_name,
                package_name=final_package_name,
                app_key=app_key,
                app_id=app_id,
                cp_id=cp_id,
                full_company_name=full_name,
                media_id=media_id,
                ad_list=ad_list,
            )
            generated_files.append(txt_path)
            _log(f"  ✓ 渠广文件已生成: {os.path.basename(txt_path)}")
            _log(f"{'='*50}")

        _log(f"\n{'='*50}")
        _log(f"全部完成，共生成 {len(generated_files)} 个渠广文件")
        _log(f"输出目录: {output_dir}")
        
        return generated_files
